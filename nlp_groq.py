"""
nlp_groq.py -- Analisis sentimen pakai Groq (Llama 3.3 70B, GRATIS)
====================================================================
Setup:
    pip install openai openpyxl

Cara pakai:
    python nlp_groq.py

API key gratis di: https://console.groq.com/keys
"""

import json, os, re, time, sys, io, collections, math
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

XLSX_PATH  = os.path.join(os.path.dirname(__file__), "ANGKET PSB.xlsx")
DATA_DIR   = os.path.join(os.path.dirname(__file__), "data", "survey_data")
BATCH_SIZE = 10       # Batch lebih kecil (10 teks) agar aman dari TPM limit
DELAY_SEC  = 1.5      # Jeda singkat antar request
MODEL_NAME = "llama-3.1-8b-instant" # Model dengan limit 30,000 TPM (5x lipat lebih besar dari 70B)

TOPIC_KEYWORDS = {
    "Aplikasi & Teknologi PSB": [
        "aplikasi","psb","online","login","wa","whatsapp","website","sistem",
        "notifikasi","upload","fitur","tombol","submit","simpan","admission",
        "teknologi","akses","url","verifikasi","form","daftar","pendaftaran",
        "digital","internet"
    ],
    "Pembayaran & Biaya": [
        "bayar","pembayaran","uang","biaya","spp","up","cicilan","dp",
        "kredit","cc","transfer","nominal","harga","mahal","terjangkau",
        "keringanan","diskon","potongan","lunas","finansial","ekonomi",
        "sibling","pengembangan"
    ],
    "Fasilitas & Infrastruktur": [
        "fasilitas","gedung","ruang","kelas","laboratorium","kantin","toilet",
        "kamar mandi","parkir","lapangan","sarana","prasarana","bangunan",
        "kebersihan","lingkungan","ekskul","ekstrakurikuler"
    ],
    "Pelayanan & Responsivitas": [
        "pelayanan","respon","response","admin","staff","komunikasi","informasi",
        "layanan","cepat","lambat","ramah","tata usaha","offline"
    ],
    "Kurikulum & Akademik": [
        "kurikulum","akademik","pelajaran","belajar","nilai","ujian","tugas",
        "guru","pengajar","inovasi","mutu","kualitas","prestasi","lomba",
        "olimpiade","bahasa","inggris","mandarin","universitas"
    ],
    "Karakter & Kedisiplinan": [
        "karakter","disiplin","sopan","santun","moral","etika","kerohanian",
        "kristen","ibadah","bullying","perilaku","sikap"
    ],
}

SKIP_SET = {
    "-","--","---",".","..","...","tidak ada","tdk ada","blm ada",
    "belum ada masukan","n/a","na","nihil","tidak ada masukan",
    "tidak ada saran","tidak ada komentar","no comment","none",""
}

STOPWORDS = set([
    "yang","dan","di","ke","dari","ini","itu","untuk","dengan","pada","ada",
    "tidak","bisa","sudah","akan","lebih","juga","saya","kami","kita","anak",
    "orang","tua","sekolah","petra","pppk","bagi","agar","jika","kalau","saat",
    "sangat","masih","nya","pun","lah","kah","deh","sih","ya","yg","jd","sy",
    "dlm","dg","lg","sdh","tdk","bs","nih","utk","krn","spy","supaya","dalam",
    "oleh","jadi","baik","semua","terus","hanya","saja","perlu","mohon",
    "harap","semoga","mungkin","tolong"
])


def load_xlsx():
    try:
        import openpyxl
    except ImportError:
        print("  [!] Jalankan: pip install openpyxl"); sys.exit(1)

    print(f"  Membaca {XLSX_PATH} ...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value) if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    idx_saran = headers.index("Saran dan masukan")
    idx_org   = headers.index("organization_code")    if "organization_code"    in headers else None
    idx_kelas = headers.index("classroom_level_code") if "classroom_level_code" in headers else None

    rows = []
    total = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        total = i
        text = str(row[idx_saran]).strip() if row[idx_saran] else ""
        if text.lower() in SKIP_SET or len(text) <= 3:
            continue
        rows.append({
            "id":      i,
            "text":    text,
            "sekolah": str(row[idx_org])   if idx_org   is not None and row[idx_org]   else "?",
            "jenjang": str(row[idx_kelas]) if idx_kelas is not None and row[idx_kelas] else "?",
        })

    print(f"  Ditemukan {len(rows)} entri bermakna dari {total} baris total.\n")
    return rows


def classify_topic(text):
    text_lower = text.lower()
    best, best_n = "Lainnya", 0
    for topic, kws in TOPIC_KEYWORDS.items():
        n = 0
        for kw in kws:
            if re.search(rf"\b{re.escape(kw)}\b", text_lower):
                n += 1
        if n > best_n:
            best_n, best = n, topic
    return best


def count_words(entries):
    counter = collections.Counter()
    for e in entries:
        tokens = re.findall(r"[a-zA-Z]{3,}", e["text"].lower())
        for t in tokens:
            if t not in STOPWORDS:
                counter[t] += 1
    return counter


def batch_sentiment(client, texts):
    numbered = "\n".join(f"{i+1}. {t[:300]}" for i, t in enumerate(texts))
    prompt = f"""Kamu adalah analis sentimen survei sekolah Bahasa Indonesia.
Klasifikasikan setiap feedback ke salah satu:
- Positif         : puas, pujian, apresiasi, harapan positif, dukungan (termasuk kalimat singkat seperti "sudah baik", "sudah bagus", "ok", "all good")
- Saran Perbaikan : ada kritik, keluhan, permintaan perbaikan, saran konkret
- Netral          : tidak ada isi/masukan yang jelas, placeholder, hanya kalimat "tidak ada" atau "tidak ada saran" saja tanpa ada sentimen pujian/kritik

Aturan penting:
- "Semoga Petra makin maju" = Positif (harapan positif)
- "Pertahankan kualitas" = Positif (bukan saran negatif)
- "tidak ada" saja atau "tidak ada saran" = Netral
- "sudah baik", "sudah bagus", "ok", "all good" = Positif
- Ada permintaan/keluhan spesifik = Saran Perbaikan

BALAS HANYA JSON array {len(texts)} elemen, tanpa penjelasan lain:
["label1", "label2", ...]

Feedback:
{numbered}
"""
    resp = client.chat.completions.create(
        model=MODEL_NAME,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.1,
        max_tokens=500,
    )
    raw = resp.choices[0].message.content.strip()

    match = re.search(r"\[.*?\]", raw, re.DOTALL)
    if not match:
        print(f"\n    [!] Parse gagal. Raw: {raw[:80]}")
        return ["Netral"] * len(texts)

    labels = json.loads(match.group())
    valid  = {"Positif", "Saran Perbaikan", "Netral"}
    result = []
    for lbl in labels:
        lbl = lbl.strip()
        if lbl in valid:
            result.append(lbl)
        elif "positif" in lbl.lower():
            result.append("Positif")
        elif any(x in lbl.lower() for x in ["saran","perbaikan","negatif"]):
            result.append("Saran Perbaikan")
        else:
            result.append("Netral")

    while len(result) < len(texts): result.append("Netral")
    return result[:len(texts)]


def build_outputs(entries):
    counts = collections.Counter(e["sentiment"] for e in entries)
    total  = len(entries)
    colors = {"Positif":"#10b981","Saran Perbaikan":"#02C5BE","Netral":"#64748b"}

    summary = [
        {"label": lbl, "count": counts[lbl],
         "percentage": round(counts[lbl]/total*100,1), "color": colors[lbl]}
        for lbl in ["Positif","Saran Perbaikan","Netral"]
    ]
    samples = {lbl: [] for lbl in ["Positif","Saran Perbaikan","Netral"]}
    for e in entries:
        if len(samples[e["sentiment"]]) < 15:
            samples[e["sentiment"]].append(e["text"])

    all_comments = [
        {"id":e["id"],"text":e["text"],"jenjang":e["jenjang"],"sekolah":e["sekolah"],
         "sentiment_raw":e["sentiment"],"sentiment":e["sentiment"],"topic":e["topic"]}
        for e in entries
    ]

    sentimen_json = {"total":total,"summary":summary,"samples":samples,"all_comments":all_comments}
    lengkap_json  = {"total":total,"data":all_comments}

    topic_map = collections.defaultdict(list)
    for e in entries: topic_map[e["topic"]].append(e["text"])
    topics_json = {
        "total": total,
        "topics": [
            {"topic":t,"count":len(v),"percentage":round(len(v)/total*100,1),"samples":v[:10]}
            for t,v in sorted(topic_map.items(), key=lambda x:-len(x[1]))
        ]
    }

    wc = count_words(entries)
    freq_json = {"total_tokens":sum(wc.values()),"unique_tokens":len(wc),
                 "top_words":[{"word":w,"count":c} for w,c in wc.most_common(100)]}

    saran_json = {"total_valid":total,"sample_saran":[e["text"] for e in entries[:20]]}

    return sentimen_json, lengkap_json, topics_json, freq_json, saran_json


def save_json(data, filename):
    path = os.path.join(DATA_DIR, filename)
    with open(path,"w",encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  Saved: {filename}")


def main():
    print("\n" + "="*60)
    print("  NLP SENTIMEN - Saran & Masukan PSB")
    print("  Powered by Groq (Llama 3.3 70B) - GRATIS")
    print("="*60 + "\n")

    try:
        from openai import OpenAI
    except ImportError:
        print("  [!] Jalankan: pip install openai"); sys.exit(1)

    api_key = input("  Masukkan Groq API key (https://console.groq.com/keys): ").strip()
    if not api_key:
        print("  [!] API key kosong."); sys.exit(1)

    client = OpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1")

    # Test koneksi
    print("  Mengetes koneksi ke Groq ...")
    try:
        test = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[{"role":"user","content":"Balas hanya: OK"}],
            max_tokens=5,
        )
        print(f"  Koneksi OK! Model: {MODEL_NAME}\n")
    except Exception as ex:
        print(f"  [!] Gagal koneksi: {ex}")
        sys.exit(1)

    entries = load_xlsx()
    texts   = [e["text"] for e in entries]
    labels  = []
    n_batch = math.ceil(len(texts) / BATCH_SIZE)

    print(f"  Memproses {len(texts)} teks dalam {n_batch} batch ...")
    print(f"  Estimasi waktu: ~{n_batch*3:.0f}-{n_batch*5:.0f} detik\n")

    for i in range(0, len(texts), BATCH_SIZE):
        batch    = texts[i:i+BATCH_SIZE]
        batch_no = i // BATCH_SIZE + 1
        print(f"  Batch {batch_no}/{n_batch} ({len(batch)} teks) ...", end=" ", flush=True)

        success = False
        for attempt in range(3):
            try:
                bl = batch_sentiment(client, batch)
                labels.extend(bl)
                p = bl.count("Positif"); s = bl.count("Saran Perbaikan"); n = bl.count("Netral")
                print(f"OK  (P:{p} S:{s} N:{n})")
                success = True
                break
            except Exception as ex:
                err_str = str(ex)
                if "429" in err_str or "rate" in err_str.lower():
                    wait = (attempt + 1) * 10
                    print(f"rate limit, tunggu {wait}s ...", end=" ", flush=True)
                    time.sleep(wait)
                else:
                    print(f"ERROR: {err_str[:100]}")
                    break

        if not success:
            print("skip (Netral)")
            labels.extend(["Netral"] * len(batch))

        if i + BATCH_SIZE < len(texts):
            time.sleep(DELAY_SEC)

    for e, lbl in zip(entries, labels):
        e["sentiment"] = lbl
        e["topic"]     = classify_topic(e["text"])

    print(f"\n  Distribusi sentimen:")
    c = collections.Counter(e["sentiment"] for e in entries)
    for lbl in ["Positif","Saran Perbaikan","Netral"]:
        print(f"    {lbl:<22}: {c[lbl]:>4}  ({c[lbl]/len(entries)*100:.1f}%)")

    print("\n  Menyimpan JSON ...\n")
    s,l,t,w,sm = build_outputs(entries)
    save_json(s,  "sentimen_saran.json")
    save_json(l,  "saran_masukan_lengkap.json")
    save_json(t,  "topik_saran.json")
    save_json(w,  "word_frequency.json")
    save_json(sm, "saran_masukan.json")

    print("\n  DONE! Refresh dashboard untuk lihat hasilnya.")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()
